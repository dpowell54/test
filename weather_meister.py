from __future__ import annotations

import json
import os
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from typing import Any

from eastern_time import now_eastern


@dataclass(frozen=True)
class WeatherMetrics:
    location: str
    observation_time: datetime
    temperature_c: float | None
    wind_kph: float | None
    humidity_percent: float | None
    condition: str | None
    raw_payload: dict[str, Any]


def _coerce_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _find_first(payload: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in payload:
            return payload[key]
    return None


def _extract_metrics(payload: dict[str, Any], location: str) -> WeatherMetrics:
    current = payload.get("current") if isinstance(payload.get("current"), dict) else payload
    temperature = _find_first(current, "temperature_c", "temp_c", "temperature", "temp")
    wind = _find_first(current, "wind_kph", "wind_speed_kph", "wind_speed", "wind")
    humidity = _find_first(current, "humidity_percent", "humidity", "relative_humidity")
    condition = _find_first(current, "condition", "summary", "weather")
    observed_at = _find_first(current, "observation_time", "observed_at", "timestamp")
    if observed_at is None:
        observed_time = now_eastern()
    else:
        try:
            observed_time = datetime.fromisoformat(str(observed_at))
        except ValueError:
            observed_time = now_eastern()

    return WeatherMetrics(
        location=location,
        observation_time=observed_time,
        temperature_c=_coerce_float(temperature),
        wind_kph=_coerce_float(wind),
        humidity_percent=_coerce_float(humidity),
        condition=str(condition) if condition is not None else None,
        raw_payload=payload,
    )


def fetch_weather(
    location: str,
    api_key: str | None = None,
    base_url: str | None = None,
    timeout: int = 15,
) -> WeatherMetrics:
    """Fetch weather data from Weather Meister and normalize it."""
    api_key = api_key or os.environ.get("WEATHER_MEISTER_API_KEY")
    if not api_key:
        raise ValueError("WEATHER_MEISTER_API_KEY is required.")
    base_url = base_url or os.environ.get("WEATHER_MEISTER_BASE_URL")
    if not base_url:
        raise ValueError("WEATHER_MEISTER_BASE_URL is required.")

    query = urllib.parse.urlencode({"location": location, "api_key": api_key})
    url = f"{base_url.rstrip('/')}/weather?{query}"
    request = urllib.request.Request(url, headers={"Accept": "application/json"})

    with urllib.request.urlopen(request, timeout=timeout) as response:
        payload = json.loads(response.read().decode("utf-8"))

    if not isinstance(payload, dict):
        raise ValueError("Unexpected response format from Weather Meister.")
    return _extract_metrics(payload, location)


def format_weather(metrics: WeatherMetrics) -> str:
    parts = [
        f"Location: {metrics.location}",
        f"Observed: {metrics.observation_time.isoformat()}",
    ]
    if metrics.temperature_c is not None:
        parts.append(f"Temp: {metrics.temperature_c:.1f}°C")
    if metrics.wind_kph is not None:
        parts.append(f"Wind: {metrics.wind_kph:.1f} kph")
    if metrics.humidity_percent is not None:
        parts.append(f"Humidity: {metrics.humidity_percent:.1f}%")
    if metrics.condition:
        parts.append(f"Condition: {metrics.condition}")
    return " | ".join(parts)


def assess_risk(
    metrics: WeatherMetrics,
    *,
    temperature_amber: float | None = None,
    temperature_red: float | None = None,
    wind_amber: float | None = None,
    wind_red: float | None = None,
    humidity_amber: float | None = None,
    humidity_red: float | None = None,
) -> str:
    """Return green/amber/red based on thresholds."""
    def _env_float(name: str) -> float | None:
        value = os.environ.get(name)
        return _coerce_float(value)

    temperature_amber = temperature_amber or _env_float("WEATHER_RISK_TEMPERATURE_AMBER")
    temperature_red = temperature_red or _env_float("WEATHER_RISK_TEMPERATURE_RED")
    wind_amber = wind_amber or _env_float("WEATHER_RISK_WIND_AMBER")
    wind_red = wind_red or _env_float("WEATHER_RISK_WIND_RED")
    humidity_amber = humidity_amber or _env_float("WEATHER_RISK_HUMIDITY_AMBER")
    humidity_red = humidity_red or _env_float("WEATHER_RISK_HUMIDITY_RED")

    checks = [
        (metrics.temperature_c, temperature_amber, temperature_red),
        (metrics.wind_kph, wind_amber, wind_red),
        (metrics.humidity_percent, humidity_amber, humidity_red),
    ]

    for value, amber, red in checks:
        if value is None or red is None:
            continue
        if value >= red:
            return "red"
    for value, amber, _red in checks:
        if value is None or amber is None:
            continue
        if value >= amber:
            return "amber"
    return "green"


def write_weather_to_excel(
    path: str,
    metrics: WeatherMetrics,
    risk_level: str,
    sheet_name: str = "Weather",
) -> None:
    """Write the weather metrics to an Excel workbook (xlsx) or CSV."""
    if path.lower().endswith(".csv"):
        _write_weather_csv(path, metrics, risk_level)
        return

    import importlib

    openpyxl_spec = importlib.util.find_spec("openpyxl")
    if openpyxl_spec is None:
        raise RuntimeError(
            "openpyxl is required for .xlsx output. Install it or use a .csv path."
        )

    from openpyxl import Workbook, load_workbook

    if os.path.exists(path):
        workbook = load_workbook(path)
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.append(
            [
                "Observed (Eastern)",
                "Location",
                "Temp (C)",
                "Wind (kph)",
                "Humidity (%)",
                "Condition",
                "Risk",
            ]
        )

    sheet.append(
        [
            metrics.observation_time.isoformat(),
            metrics.location,
            metrics.temperature_c,
            metrics.wind_kph,
            metrics.humidity_percent,
            metrics.condition,
            risk_level,
        ]
    )
    workbook.save(path)


def _write_weather_csv(path: str, metrics: WeatherMetrics, risk_level: str) -> None:
    import csv

    file_exists = os.path.exists(path)
    with open(path, "a", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        if not file_exists:
            writer.writerow(
                [
                    "Observed (Eastern)",
                    "Location",
                    "Temp (C)",
                    "Wind (kph)",
                    "Humidity (%)",
                    "Condition",
                    "Risk",
                ]
            )
        writer.writerow(
            [
                metrics.observation_time.isoformat(),
                metrics.location,
                metrics.temperature_c,
                metrics.wind_kph,
                metrics.humidity_percent,
                metrics.condition,
                risk_level,
            ]
        )
